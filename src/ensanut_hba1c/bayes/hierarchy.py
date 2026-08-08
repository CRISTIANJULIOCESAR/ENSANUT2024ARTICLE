"""Fast descriptive Bayes evidence summaries by ENSANUT dictionary hierarchy.

This module intentionally does *not* compute predictive performance metrics.
There are no ROC curves, AUC values, thresholds, confusion matrices, accuracy,
F1, calibration statistics, train/test splits, or cross-validation.

For each primary or secondary dictionary group, predictors are encoded exactly
once. All SHAP-derived Leiden microclusters are then evaluated simultaneously
with vectorized contingency-table calculations.

For every group × cluster combination, the module reports:

- sum of positive Bayes log-likelihood-ratio scores;
- absolute and signed sum of negative scores;
- net score sum;
- counts of positive, negative, and neutral rules;
- raw positive and negative score sums for audit;
- category-frequency- and missingness-weighted evidence;
- weighted evidence normalized per effective variable for fair comparison
  between questionnaire sections of different sizes.

The analysis is descriptive and uses the complete cohort.
"""

from __future__ import annotations

import json
import math
import re
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from tqdm.auto import tqdm
except ImportError:  # pragma: no cover
    tqdm = None

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from .analysis import BIN_SEP, codificar_predictores_bayes, normalizar_codigo

NO_SUBLEVEL = "(sin subnivel)"
ANALYSIS_SCOPE = "descriptive_weighted_signed_bayes_evidence_full_dataset"


# Identification columns shared by the principal table and the internal
# calculation-audit workbooks.
RULE_IDENTIFIER_COLUMNS = [
    "level",
    "group",
    "primary_group",
    "secondary_group",
    "cluster",
    "var",
    "var_alias",
    "description",
    "pregunta",
    "tipo_variable",
    "categoria_o_rango",
    "valor_crudo",
]

# Principal publication-oriented table.  It contains final evidence measures
# and coverage, but not every intermediate mathematical component.
MASTER_RULE_COLUMNS = RULE_IDENTIFIER_COLUMNS + [
    "Nx",
    "n_cluster",
    "n_cluster_x",
    "coverage_in_cluster",
    "p_c_given_x",
    "epsilon",
    "score",
]

# One row-wise audit table containing every count, coverage, score and epsilon
# operation.  All intermediate components are exported together to one Excel
# workbook so each rule can be reproduced from left to right in the same row.
INTERNAL_OPERATION_COLUMNS = RULE_IDENTIFIER_COLUMNS + [
    "alpha",
    "min_cases",
    "N",
    "n_cluster",
    "n_rest",
    "Nx",
    "n_cluster_x",
    "n_rest_x",
    "coverage_in_cluster",
    "coverage_in_rest",
    "prior_cluster",
    "p_x_given_cluster_numerator",
    "p_x_given_cluster_denominator",
    "p_x_given_cluster",
    "p_x_given_rest_numerator",
    "p_x_given_rest_denominator",
    "p_x_given_rest",
    "likelihood_ratio",
    "raw_score",
    "score_aplicado",
    "score",
    "p_c_given_x_numerator",
    "p_c_given_x_denominator",
    "p_c_given_x",
    "epsilon_difference",
    "epsilon_numerator",
    "epsilon_denominator",
    "epsilon",
]

# Subsets remain available in memory for backward-compatible programmatic use,
# but they are no longer written to separate Excel files.
COUNT_OPERATION_COLUMNS = RULE_IDENTIFIER_COLUMNS + [
    "alpha", "min_cases", "N", "n_cluster", "n_rest", "Nx",
    "n_cluster_x", "n_rest_x", "coverage_in_cluster", "coverage_in_rest",
]
SCORE_OPERATION_COLUMNS = RULE_IDENTIFIER_COLUMNS + [
    "alpha", "min_cases", "Nx", "n_cluster", "n_rest", "n_cluster_x",
    "n_rest_x", "p_x_given_cluster_numerator",
    "p_x_given_cluster_denominator", "p_x_given_cluster",
    "p_x_given_rest_numerator", "p_x_given_rest_denominator",
    "p_x_given_rest", "likelihood_ratio", "raw_score",
    "score_aplicado", "score",
]
EPSILON_OPERATION_COLUMNS = RULE_IDENTIFIER_COLUMNS + [
    "alpha", "N", "Nx", "n_cluster", "n_cluster_x", "prior_cluster",
    "p_c_given_x_numerator", "p_c_given_x_denominator", "p_c_given_x",
    "epsilon_difference", "epsilon_numerator", "epsilon_denominator",
    "epsilon",
]


def _first_metadata_text(metadata: dict[str, Any], *keys: str) -> str:
    """Return the first non-empty dictionary value among candidate keys."""
    for key in keys:
        value = metadata.get(key)
        if value is None or (isinstance(value, float) and np.isnan(value)):
            continue
        text = re.sub(r"\s+", " ", str(value).strip())
        if text and text.lower() not in {"nan", "none", "<na>"}:
            return text
    return ""


def _questionnaire_groups_from_metadata(
    metadata: dict[str, Any],
    *,
    analysis_level: str,
    analysis_group: str,
) -> tuple[str, str]:
    """Resolve the primary and secondary questionnaire groups for one variable.

    The original ENSANUT dictionary normally stores both levels in
    ``subcategoria`` separated at the first period. Explicit hierarchy columns
    are preferred when available. The analysis group is used only as a safe
    fallback.
    """
    primary = _first_metadata_text(
        metadata,
        "nivel_primario",
        "primary_group",
        "primary_section",
        "grupo_primario",
    )
    secondary = _first_metadata_text(
        metadata,
        "nivel_secundario",
        "secondary_group",
        "secondary_section",
        "grupo_secundario",
    )

    route = _first_metadata_text(
        metadata,
        "ruta_jerarquica",
        "subcategoria",
        "subcategory",
    )
    if route:
        if " > " in route:
            route_primary, route_secondary = route.split(" > ", 1)
        elif "." in route:
            route_primary, route_secondary = route.split(".", 1)
        else:
            route_primary, route_secondary = route, ""
        primary = primary or route_primary.strip()
        secondary = secondary or route_secondary.strip()

    group_text = str(analysis_group).strip()
    if analysis_level == "secondary":
        if " > " in group_text:
            fallback_primary, fallback_secondary = group_text.split(" > ", 1)
        elif "." in group_text:
            fallback_primary, fallback_secondary = group_text.split(".", 1)
        else:
            fallback_primary, fallback_secondary = group_text, ""
        primary = primary or fallback_primary.strip()
        secondary = secondary or fallback_secondary.strip()
    else:
        primary = primary or group_text

    return primary, secondary or NO_SUBLEVEL


@dataclass(frozen=True)
class SignedEvidenceConfig:
    n_bins_features: int = 10
    alpha: float = 1.0
    min_cases: int = 20
    epsilon_threshold: float = 1.96
    force_exact_feature_bins: bool = False
    max_inferred_categories: int = 50
    minimum_variables_per_group: int = 2
    top_rules_per_cluster: int = 10
    publication_dpi: int = 600
    show_progress_bar: bool = True
    progress_leave: bool = True
    progress_mininterval: float = 0.25
    save_group_figures: bool = False
    save_top_rule_tables: bool = True
    save_complex_heatmaps: bool = True
    complex_heatmap_robust_percentile: float = 95.0
    export_min_nx: int = 40
    export_min_n_cluster_x: int = 40
    export_min_score: float = 0.5


def _normalize_text(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    return text or None


def _safe_name(value: Any, max_length: int = 120) -> str:
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("_") or "group"
    return text[:max_length]


def _cluster_sort_key(value: Any) -> tuple[int, float | str]:
    try:
        return 0, float(value)
    except (TypeError, ValueError):
        return 1, str(value)


def parse_dictionary_hierarchy(
    dictionary_df: pd.DataFrame,
    *,
    subcategory_column: str = "subcategoria",
    variable_column: str = "var",
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, list[str]], dict[str, list[str]]]:
    """Split ``subcategoria`` at the first dot into primary and secondary levels."""

    if subcategory_column not in dictionary_df.columns:
        raise KeyError(f"Dictionary column not found: {subcategory_column}")
    if variable_column not in dictionary_df.columns:
        raise KeyError(f"Dictionary variable column not found: {variable_column}")

    out = dictionary_df.loc[:, ~dictionary_df.columns.duplicated()].copy()
    out["subcategoria_original"] = out[subcategory_column].astype("string")
    out["subcategoria_limpia"] = out["subcategoria_original"].map(_normalize_text)

    primary_values: list[str | None] = []
    secondary_values: list[str] = []
    for value in out["subcategoria_limpia"]:
        if value is None or pd.isna(value):
            primary_values.append(None)
            secondary_values.append(NO_SUBLEVEL)
            continue
        parts = str(value).split(".", 1)
        primary = _normalize_text(parts[0])
        secondary = _normalize_text(parts[1]) if len(parts) == 2 else None
        primary_values.append(primary)
        secondary_values.append(secondary or NO_SUBLEVEL)

    out["nivel_primario"] = primary_values
    out["nivel_secundario"] = secondary_values
    out["ruta_jerarquica"] = np.where(
        out["nivel_secundario"].eq(NO_SUBLEVEL),
        out["nivel_primario"],
        out["nivel_primario"].astype("string") + " > " + out["nivel_secundario"].astype("string"),
    )

    valid = out.dropna(subset=[variable_column, "nivel_primario"]).copy()
    valid[variable_column] = valid[variable_column].astype(str)

    summary = (
        valid.groupby(
            ["nivel_primario", "nivel_secundario", "ruta_jerarquica"],
            dropna=False,
            observed=True,
        )
        .agg(
            n_filas=(variable_column, "size"),
            n_variables_unicas=(variable_column, "nunique"),
        )
        .reset_index()
        .sort_values(["nivel_primario", "nivel_secundario"], kind="stable")
        .reset_index(drop=True)
    )

    primary_groups = {
        str(primary): sorted(group[variable_column].dropna().astype(str).unique().tolist())
        for primary, group in valid.groupby("nivel_primario", observed=True)
    }
    secondary_groups = {
        str(route): sorted(group[variable_column].dropna().astype(str).unique().tolist())
        for route, group in valid.groupby("ruta_jerarquica", observed=True)
    }
    return out, summary, primary_groups, secondary_groups


def hierarchy_tree_text(summary: pd.DataFrame) -> str:
    lines: list[str] = []
    for primary in summary["nivel_primario"].drop_duplicates().tolist():
        primary_rows = summary.loc[summary["nivel_primario"].eq(primary)]
        lines.append(f"■ {primary}")
        secondaries = primary_rows["nivel_secundario"].tolist()
        for index, secondary in enumerate(secondaries):
            row = primary_rows.iloc[index]
            branch = "└──" if index == len(secondaries) - 1 else "├──"
            lines.append(
                f"  {branch} {secondary} "
                f"({int(row['n_variables_unicas'])} variables; {int(row['n_filas'])} rows)"
            )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _dictionary_metadata_map(dictionary_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    deduplicated = dictionary_df.drop_duplicates(subset=["var"], keep="first")
    return {
        str(row["var"]): row.to_dict()
        for _, row in deduplicated.iterrows()
        if pd.notna(row.get("var"))
    }


def _vectorized_scores_for_all_clusters(
    X_bin: pd.DataFrame,
    cluster_series: pd.Series,
    *,
    alpha: float,
    min_cases: int,
) -> dict[str, Any]:
    """Calculate all cluster-specific Bayes rule scores in one matrix pass."""

    clusters = sorted(cluster_series.dropna().unique(), key=_cluster_sort_key)
    if len(clusters) < 2:
        raise ValueError("At least two clusters are required.")

    cluster_codes = [normalizar_codigo(cluster) for cluster in clusters]
    normalized = cluster_series.map(normalizar_codigo)
    Y = np.column_stack([
        normalized.eq(code).to_numpy(dtype=np.uint8)
        for code in cluster_codes
    ])

    X = X_bin.to_numpy(dtype=np.uint8, copy=False)
    N = int(X.shape[0])
    Nx = X.sum(axis=0, dtype=np.int64)
    Nc = Y.sum(axis=0, dtype=np.int64)
    Nnc = N - Nc

    # P × K contingency counts for every binary predictor and cluster.
    nCx = X.T.astype(np.int64, copy=False) @ Y.astype(np.int64, copy=False)
    nnCx = Nx[:, None] - nCx

    # Internal operations for the Bayes score.  Every numerator and denominator
    # is retained so the exported audit workbooks can reproduce the result row
    # by row without hidden steps.
    p_x_given_c_numerator = nCx + alpha
    p_x_given_c_denominator = Nc[None, :] + 2 * alpha
    p_x_given_c = p_x_given_c_numerator / p_x_given_c_denominator

    p_x_given_nc_numerator = nnCx + alpha
    p_x_given_nc_denominator = Nnc[None, :] + 2 * alpha
    p_x_given_nc = p_x_given_nc_numerator / p_x_given_nc_denominator

    likelihood_ratio = p_x_given_c / p_x_given_nc
    raw_score = np.log(likelihood_ratio)
    applied = (Nx >= min_cases).astype(np.uint8)
    score = np.where(applied[:, None] == 1, raw_score, 0.0)

    # Probability of cluster membership among participants who satisfy the rule.
    prior = Nc.astype(float) / max(N, 1)
    p_c_given_x_numerator = nCx + alpha
    p_c_given_x_denominator = Nx[:, None] + 2 * alpha
    p_c_given_x = p_c_given_x_numerator / p_c_given_x_denominator

    # Standardized epsilon statistic used to identify bins whose cluster
    # frequency departs from the global cluster prevalence.
    epsilon_difference = p_c_given_x - prior[None, :]
    epsilon_numerator = Nx[:, None] * epsilon_difference
    epsilon_denominator = np.sqrt(
        Nx[:, None] * prior[None, :] * (1.0 - prior[None, :])
    )
    epsilon = np.divide(
        epsilon_numerator,
        epsilon_denominator,
        out=np.full_like(p_c_given_x, np.nan, dtype=float),
        where=epsilon_denominator > 0,
    )

    return {
        "clusters": clusters,
        "cluster_codes": cluster_codes,
        "Y": Y,
        "alpha": float(alpha),
        "min_cases": int(min_cases),
        "N": N,
        "Nx": Nx,
        "Nc": Nc,
        "Nnc": Nnc,
        "nCx": nCx,
        "nnCx": nnCx,
        "prior": prior,
        "p_x_given_c_numerator": p_x_given_c_numerator,
        "p_x_given_c_denominator": p_x_given_c_denominator,
        "p_x_given_c": p_x_given_c,
        "p_x_given_nc_numerator": p_x_given_nc_numerator,
        "p_x_given_nc_denominator": p_x_given_nc_denominator,
        "p_x_given_nc": p_x_given_nc,
        "p_c_given_x_numerator": p_c_given_x_numerator,
        "p_c_given_x_denominator": p_c_given_x_denominator,
        "p_c_given_x": p_c_given_x,
        "likelihood_ratio": likelihood_ratio,
        "raw_score": raw_score,
        "applied": applied,
        "score": score,
        "epsilon_difference": epsilon_difference,
        "epsilon_numerator": epsilon_numerator,
        "epsilon_denominator": epsilon_denominator,
        "epsilon": epsilon,
    }


def _build_rule_detail_tables(
    *,
    encoded: dict[str, Any],
    score_data: dict[str, Any],
    dictionary_df: pd.DataFrame,
    level: str,
    group: str,
    top_n: int,
    epsilon_threshold: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Create compact top-positive/top-negative rule tables and a rule audit."""

    X_columns = encoded["X_bin"].columns.astype(str).tolist()
    label_map = encoded["label_map"]
    type_map = encoded["type_map"]
    metadata_map = _dictionary_metadata_map(dictionary_df)

    variables: list[str] = []
    raw_values: list[str] = []
    labels: list[str] = []
    aliases: list[str] = []
    descriptions: list[str] = []
    questions: list[str] = []
    primary_groups: list[str] = []
    secondary_groups: list[str] = []
    variable_types: list[str] = []

    for column in X_columns:
        if BIN_SEP in column:
            variable, raw_value = column.split(BIN_SEP, 1)
        else:
            variable, raw_value = column, "1"
        metadata = metadata_map.get(variable, {})
        variables.append(variable)
        raw_values.append(raw_value)
        labels.append(label_map.get((variable, raw_value), raw_value))
        aliases.append(_first_metadata_text(metadata, "var_alias", "alias") or variable)
        description = _first_metadata_text(
            metadata,
            "description",
            "descripcion",
            "variable_description",
        ) or variable
        question = _first_metadata_text(
            metadata,
            "pregunta",
            "question",
            "texto_pregunta",
            "descripcion_pregunta",
            "item_text",
            "label",
        ) or description
        primary_group, secondary_group = _questionnaire_groups_from_metadata(
            metadata,
            analysis_level=level,
            analysis_group=group,
        )
        descriptions.append(description)
        questions.append(question)
        primary_groups.append(primary_group)
        secondary_groups.append(secondary_group)
        variable_types.append(type_map.get(variable, "unknown"))

    all_rows: list[pd.DataFrame] = []
    positive_rows: list[pd.DataFrame] = []
    negative_rows: list[pd.DataFrame] = []

    for cluster_index, cluster in enumerate(score_data["clusters"]):
        frame = pd.DataFrame({
            "level": level,
            "group": group,
            "primary_group": primary_groups,
            "secondary_group": secondary_groups,
            "cluster": cluster,
            "var": variables,
            "var_alias": aliases,
            "description": descriptions,
            "pregunta": questions,
            "tipo_variable": variable_types,
            "categoria_o_rango": labels,
            "valor_crudo": raw_values,
            "alpha": score_data["alpha"],
            "min_cases": score_data["min_cases"],
            "N": score_data["N"],
            "Nc": score_data["Nc"][cluster_index],
            "Nnc": score_data["Nnc"][cluster_index],
            "n_cluster": score_data["Nc"][cluster_index],
            "n_rest": score_data["Nnc"][cluster_index],
            "prior": score_data["prior"][cluster_index],
            "prior_cluster": score_data["prior"][cluster_index],
            "Nx": score_data["Nx"],
            "n_cluster_x": score_data["nCx"][:, cluster_index],
            "coverage_in_cluster": (
                score_data["nCx"][:, cluster_index]
                / max(int(score_data["Nc"][cluster_index]), 1)
            ),
            "n_rest_x": score_data["nnCx"][:, cluster_index],
            "coverage_in_rest": (
                score_data["nnCx"][:, cluster_index]
                / max(int(score_data["Nnc"][cluster_index]), 1)
            ),
            "p_x_given_cluster_numerator": score_data["p_x_given_c_numerator"][:, cluster_index],
            "p_x_given_cluster_denominator": score_data["p_x_given_c_denominator"][0, cluster_index],
            "p_x_given_cluster": score_data["p_x_given_c"][:, cluster_index],
            "p_x_given_rest_numerator": score_data["p_x_given_nc_numerator"][:, cluster_index],
            "p_x_given_rest_denominator": score_data["p_x_given_nc_denominator"][0, cluster_index],
            "p_x_given_rest": score_data["p_x_given_nc"][:, cluster_index],
            "p_c_given_x_numerator": score_data["p_c_given_x_numerator"][:, cluster_index],
            "p_c_given_x_denominator": score_data["p_c_given_x_denominator"][:, 0],
            "p_c_given_x": score_data["p_c_given_x"][:, cluster_index],
            "likelihood_ratio": score_data["likelihood_ratio"][:, cluster_index],
            "raw_score": score_data["raw_score"][:, cluster_index],
            "score_aplicado": score_data["applied"],
            "score": score_data["score"][:, cluster_index],
            "epsilon_difference": score_data["epsilon_difference"][:, cluster_index],
            "epsilon_numerator": score_data["epsilon_numerator"][:, cluster_index],
            "epsilon_denominator": score_data["epsilon_denominator"][:, cluster_index],
            "epsilon": score_data["epsilon"][:, cluster_index],
        })
        all_rows.append(frame)
        positive_rows.append(
            frame.loc[frame["score"] > 0]
            .nlargest(top_n, "score")
            .reset_index(drop=True)
        )
        negative_rows.append(
            frame.loc[frame["score"] < 0]
            .nsmallest(top_n, "score")
            .reset_index(drop=True)
        )

    return (
        pd.concat(positive_rows, ignore_index=True) if positive_rows else pd.DataFrame(),
        pd.concat(negative_rows, ignore_index=True) if negative_rows else pd.DataFrame(),
        pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame(),
    )


def analyze_signed_evidence_group(
    *,
    cluster_df: pd.DataFrame,
    dictionary_df: pd.DataFrame,
    variables: Sequence[str],
    cluster_column: str,
    level: str,
    group: str,
    output_dir: str | Path,
    config: SignedEvidenceConfig,
) -> dict[str, Any]:
    """Encode one group once and summarize positive/negative evidence for all clusters."""

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    usable_variables = [
        variable for variable in variables
        if variable in cluster_df.columns and variable != cluster_column
    ]
    if len(usable_variables) < config.minimum_variables_per_group:
        raise ValueError(
            f"Group {group!r} has only {len(usable_variables)} usable variables; "
            f"minimum is {config.minimum_variables_per_group}."
        )

    group_data = cluster_df[[*usable_variables, cluster_column]].copy()
    group_dictionary = dictionary_df.loc[
        dictionary_df["var"].astype(str).isin(usable_variables)
    ].copy()

    encoded = codificar_predictores_bayes(
        group_data,
        group_dictionary,
        excluded_variables={cluster_column},
        n_bins_features=config.n_bins_features,
        force_exact_feature_bins=config.force_exact_feature_bins,
        max_categorias_inferidas=config.max_inferred_categories,
    )
    score_data = _vectorized_scores_for_all_clusters(
        encoded["X_bin"],
        group_data[cluster_column],
        alpha=config.alpha,
        min_cases=config.min_cases,
    )

    positive_scores = np.clip(score_data["score"], 0.0, None)
    negative_scores = np.clip(score_data["score"], None, 0.0)

    # Raw structural sums are retained for audit only. They are not used as
    # the main cross-section visualization because groups with more variables
    # or categories would otherwise dominate by construction.
    structural_positive = positive_scores.sum(axis=0)
    structural_negative_signed = negative_scores.sum(axis=0)
    structural_negative_abs = -structural_negative_signed
    structural_net = score_data["score"].sum(axis=0)

    # Rule frequency in the complete cohort. For a given variable, the one-hot
    # rule frequencies sum to that variable's observed-data coverage.
    rule_frequency = score_data["Nx"].astype(float) / max(score_data["N"], 1)
    weighted_positive_sum = (rule_frequency[:, None] * positive_scores).sum(axis=0)
    weighted_negative_signed = (rule_frequency[:, None] * negative_scores).sum(axis=0)
    weighted_negative_abs = -weighted_negative_signed
    weighted_net_sum = weighted_positive_sum - weighted_negative_abs

    # Each encoded variable contributes approximately one unit at complete
    # coverage, independently of whether it was represented by 2 or 10 rules.
    binary_variables = np.array([
        str(column).split(BIN_SEP, 1)[0]
        for column in encoded["X_bin"].columns
    ], dtype=object)
    encoded_variable_names = list(dict.fromkeys(binary_variables.tolist()))
    variable_coverages: dict[str, float] = {}
    for variable in encoded_variable_names:
        mask = binary_variables == variable
        valid_count = float(score_data["Nx"][mask].sum())
        variable_coverages[variable] = valid_count / max(score_data["N"], 1)

    effective_variable_count = float(sum(variable_coverages.values()))
    n_encoded_variables = int(len(encoded_variable_names))
    mean_variable_coverage = (
        effective_variable_count / n_encoded_variables
        if n_encoded_variables > 0 else np.nan
    )

    positive_evidence_per_variable = np.divide(
        weighted_positive_sum,
        effective_variable_count,
        out=np.full_like(weighted_positive_sum, np.nan, dtype=float),
        where=effective_variable_count > 0,
    )
    negative_evidence_per_variable = np.divide(
        weighted_negative_abs,
        effective_variable_count,
        out=np.full_like(weighted_negative_abs, np.nan, dtype=float),
        where=effective_variable_count > 0,
    )
    net_evidence_per_variable = (
        positive_evidence_per_variable - negative_evidence_per_variable
    )

    # Member/rest sums remain available as supplementary descriptive audits.
    member_positive = (score_data["nCx"] * positive_scores).sum(axis=0)
    member_negative_signed = (score_data["nCx"] * negative_scores).sum(axis=0)
    member_negative_abs = -member_negative_signed
    rest_positive = (score_data["nnCx"] * positive_scores).sum(axis=0)
    rest_negative_signed = (score_data["nnCx"] * negative_scores).sum(axis=0)
    rest_negative_abs = -rest_negative_signed

    rows: list[dict[str, Any]] = []
    for index, cluster in enumerate(score_data["clusters"]):
        total_abs_rule = structural_positive[index] + structural_negative_abs[index]
        rows.append({
            "analysis_scope": ANALYSIS_SCOPE,
            "level": level,
            "group": group,
            "cluster": cluster,
            "n_participants": score_data["N"],
            "n_cluster_members": int(score_data["Nc"][index]),
            "n_rest": int(score_data["Nnc"][index]),
            "n_group_variables": len(usable_variables),
            "n_encoded_variables": n_encoded_variables,
            "n_binary_rules": int(encoded["X_bin"].shape[1]),
            "effective_variable_count": effective_variable_count,
            "mean_variable_coverage": mean_variable_coverage,
            "n_positive_rules": int((score_data["score"][:, index] > 0).sum()),
            "n_negative_rules": int((score_data["score"][:, index] < 0).sum()),
            "n_neutral_rules": int((score_data["score"][:, index] == 0).sum()),
            "positive_rule_score_sum": float(structural_positive[index]),
            "negative_rule_score_sum_signed": float(structural_negative_signed[index]),
            "negative_rule_score_sum_abs": float(structural_negative_abs[index]),
            "net_rule_score_sum": float(structural_net[index]),
            "positive_rule_share": (
                float(structural_positive[index] / total_abs_rule)
                if total_abs_rule > 0 else np.nan
            ),
            "negative_rule_share": (
                float(structural_negative_abs[index] / total_abs_rule)
                if total_abs_rule > 0 else np.nan
            ),
            # Weighted quantities used by the complex heatmaps.
            "positive_evidence_weighted_sum": float(weighted_positive_sum[index]),
            "negative_evidence_weighted_sum_signed": float(weighted_negative_signed[index]),
            "negative_evidence_weighted_sum_abs": float(weighted_negative_abs[index]),
            "net_evidence_weighted_sum": float(weighted_net_sum[index]),
            "positive_evidence_per_variable": float(positive_evidence_per_variable[index]),
            "negative_evidence_per_variable": float(negative_evidence_per_variable[index]),
            "net_evidence_per_variable": float(net_evidence_per_variable[index]),
            "positive_evidence_weighted_share": (
                float(weighted_positive_sum[index] / (weighted_positive_sum[index] + weighted_negative_abs[index]))
                if (weighted_positive_sum[index] + weighted_negative_abs[index]) > 0 else np.nan
            ),
            "negative_evidence_weighted_share": (
                float(weighted_negative_abs[index] / (weighted_positive_sum[index] + weighted_negative_abs[index]))
                if (weighted_positive_sum[index] + weighted_negative_abs[index]) > 0 else np.nan
            ),
            # Backward-compatible aliases.
            "positive_cohort_contribution_sum": float(weighted_positive_sum[index]),
            "negative_cohort_contribution_sum_signed": float(weighted_negative_signed[index]),
            "negative_cohort_contribution_sum_abs": float(weighted_negative_abs[index]),
            "net_cohort_contribution_sum": float(weighted_net_sum[index]),
            "positive_cohort_share": (
                float(weighted_positive_sum[index] / (weighted_positive_sum[index] + weighted_negative_abs[index]))
                if (weighted_positive_sum[index] + weighted_negative_abs[index]) > 0 else np.nan
            ),
            "negative_cohort_share": (
                float(weighted_negative_abs[index] / (weighted_positive_sum[index] + weighted_negative_abs[index]))
                if (weighted_positive_sum[index] + weighted_negative_abs[index]) > 0 else np.nan
            ),
            "member_positive_contribution_sum": float(member_positive[index]),
            "member_negative_contribution_sum_signed": float(member_negative_signed[index]),
            "member_negative_contribution_sum_abs": float(member_negative_abs[index]),
            "member_net_contribution_sum": float(member_positive[index] - member_negative_abs[index]),
            "rest_positive_contribution_sum": float(rest_positive[index]),
            "rest_negative_contribution_sum_signed": float(rest_negative_signed[index]),
            "rest_negative_contribution_sum_abs": float(rest_negative_abs[index]),
            "rest_net_contribution_sum": float(rest_positive[index] - rest_negative_abs[index]),
            "encoding_passes": 1,
            "clusters_computed_simultaneously": len(score_data["clusters"]),
        })

    summary = pd.DataFrame(rows)
    top_positive, top_negative, all_rules = _build_rule_detail_tables(
        encoded=encoded,
        score_data=score_data,
        dictionary_df=group_dictionary,
        level=level,
        group=group,
        top_n=config.top_rules_per_cluster,
        epsilon_threshold=config.epsilon_threshold,
    )

    summary.to_csv(output_dir / "signed_evidence_summary.csv", index=False)
    encoded["encoding_diagnostics"].to_csv(
        output_dir / "encoding_diagnostics.csv", index=False
    )
    pd.DataFrame({
        "var": list(variable_coverages),
        "coverage": list(variable_coverages.values()),
    }).to_csv(output_dir / "variable_coverage.csv", index=False)
    # Complete per-bin audit requested for interpretation.
    all_rules.loc[:, MASTER_RULE_COLUMNS].to_csv(
        output_dir / "bayes_variables_bins_epsilon_score.csv", index=False
    )
    if config.save_top_rule_tables:
        top_positive.to_csv(output_dir / "top_positive_rules.csv", index=False)
        top_negative.to_csv(output_dir / "top_negative_rules.csv", index=False)

    metadata = {
        "analysis_scope": ANALYSIS_SCOPE,
        "level": level,
        "group": group,
        "n_variables_requested": len(variables),
        "n_variables_used": len(usable_variables),
        "n_encoded_variables": n_encoded_variables,
        "n_binary_rules": int(encoded["X_bin"].shape[1]),
        "effective_variable_count": effective_variable_count,
        "mean_variable_coverage": mean_variable_coverage,
        "n_clusters": len(score_data["clusters"]),
        "epsilon_threshold": float(config.epsilon_threshold),
        "encoding_passes": 1,
        "clusters_computed_simultaneously": True,
        "metrics_computed": False,
        "visualization_measure": "weighted evidence per effective variable",
    }
    (output_dir / "group_metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    figure_paths: list[Path] = []
    if config.save_group_figures:
        figure = plot_group_signed_evidence(summary, title=f"{level}: {group}")
        figure_paths = save_figure_bundle(
            figure,
            output_dir / "signed_evidence_all_clusters",
            dpi=config.publication_dpi,
        )
        plt.close(figure)

    return {
        "summary": summary,
        "top_positive_rules": top_positive,
        "top_negative_rules": top_negative,
        "all_rules": all_rules,
        "encoding_diagnostics": encoded["encoding_diagnostics"],
        "metadata": metadata,
        "figure_paths": figure_paths,
        "variables": usable_variables,
        "variable_coverage": variable_coverages,
    }


def plot_group_signed_evidence(summary: pd.DataFrame, *, title: str | None = None):
    """Weighted evidence per effective variable for all microclusters."""

    ordered = summary.copy().sort_values("cluster", key=lambda s: s.map(_cluster_sort_key))
    labels = [f"C{value}" for value in ordered["cluster"]]
    y = np.arange(len(labels))

    positive = ordered["positive_evidence_per_variable"].to_numpy(dtype=float)
    negative = -ordered["negative_evidence_per_variable"].to_numpy(dtype=float)

    fig, ax = plt.subplots(figsize=(9.0, max(4.2, 0.42 * len(labels) + 1.5)))
    ax.barh(y, negative, label="Weighted negative evidence")
    ax.barh(y, positive, label="Weighted positive evidence")
    ax.axvline(0, linewidth=0.8)
    ax.set_yticks(y, labels)
    ax.invert_yaxis()
    ax.set_xlabel("Mean Bayes evidence per effective variable")
    if title:
        ax.set_title(title)
    ax.legend(frameon=False, ncol=2, loc="upper center", bbox_to_anchor=(0.5, -0.08))
    ax.grid(axis="x", alpha=0.2)
    fig.tight_layout()
    return fig

def save_figure_bundle(figure, base_path: str | Path, *, dpi: int = 600) -> list[Path]:
    base_path = Path(base_path)
    base_path.parent.mkdir(parents=True, exist_ok=True)
    paths = [base_path.with_suffix(".png"), base_path.with_suffix(".pdf"), base_path.with_suffix(".svg")]
    figure.savefig(paths[0], dpi=dpi, bbox_inches="tight")
    figure.savefig(paths[1], bbox_inches="tight")
    figure.savefig(paths[2], bbox_inches="tight")
    return paths


def build_weighted_evidence_matrix(
    summary: pd.DataFrame,
    *,
    level: str,
    value_column: str = "net_evidence_per_variable",
) -> pd.DataFrame:
    """Pivot one weighted-evidence column to group × microcluster format."""

    subset = summary.loc[summary["level"].eq(level)].copy()
    if subset.empty:
        raise ValueError(f"No rows available for level {level!r}.")
    matrix = subset.pivot(index="group", columns="cluster", values=value_column)
    return matrix.reindex(columns=sorted(matrix.columns, key=_cluster_sort_key))


def plot_global_heatmap(
    summary: pd.DataFrame,
    *,
    level: str,
    value_column: str = "net_evidence_per_variable",
    title: str | None = None,
):
    """Compact matplotlib fallback; primary figures use PyComplexHeatmap."""

    matrix = build_weighted_evidence_matrix(
        summary, level=level, value_column=value_column
    )
    height = max(4.5, min(20.0, 0.28 * len(matrix.index) + 2.0))
    width = max(7.0, 0.7 * len(matrix.columns) + 3.5)
    fig, ax = plt.subplots(figsize=(width, height))
    values = matrix.to_numpy(dtype=float)
    if "net_" in value_column:
        limit = float(np.nanpercentile(np.abs(values), 95)) if np.isfinite(values).any() else 1.0
        limit = max(limit, np.finfo(float).eps)
        image = ax.imshow(values, aspect="auto", cmap="RdBu_r", vmin=-limit, vmax=limit)
    else:
        vmax = float(np.nanpercentile(values, 95)) if np.isfinite(values).any() else 1.0
        vmax = max(vmax, np.finfo(float).eps)
        image = ax.imshow(values, aspect="auto", vmin=0, vmax=vmax)
    ax.set_xticks(np.arange(len(matrix.columns)), [f"C{c}" for c in matrix.columns])
    ax.set_yticks(np.arange(len(matrix.index)), matrix.index)
    ax.set_xlabel("Microcluster")
    ax.set_ylabel("Questionnaire section")
    if title:
        ax.set_title(title)
    colorbar = fig.colorbar(image, ax=ax, fraction=0.025, pad=0.02)
    colorbar.set_label(value_column)
    fig.tight_layout()
    return fig, matrix

def _selected_groups(
    groups: dict[str, list[str]],
    selected: Sequence[str] | None,
) -> dict[str, list[str]]:
    if selected is None:
        return groups
    requested = {str(value) for value in selected}
    return {name: variables for name, variables in groups.items() if name in requested}


def run_hierarchical_signed_evidence_analysis(
    *,
    cluster_df: pd.DataFrame,
    dictionary_df: pd.DataFrame,
    output_dir: str | Path,
    cluster_column: str = "Cluster_SHAP",
    excluded_columns: Iterable[str] | None = None,
    run_primary_level: bool = True,
    run_secondary_level: bool = True,
    config: SignedEvidenceConfig | None = None,
    selected_primary_groups: Sequence[str] | None = None,
    selected_secondary_groups: Sequence[str] | None = None,
) -> dict[str, Any]:
    """Run the simplified positive/negative evidence analysis."""

    config = config or SignedEvidenceConfig()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    hierarchy_df, hierarchy_summary, primary_groups, secondary_groups = parse_dictionary_hierarchy(
        dictionary_df
    )
    excluded = {str(value) for value in (excluded_columns or [])}
    eligible = set(map(str, cluster_df.columns)).difference(excluded).difference({cluster_column})

    primary_groups = {
        group: sorted(set(map(str, variables)).intersection(eligible))
        for group, variables in primary_groups.items()
    }
    secondary_groups = {
        group: sorted(set(map(str, variables)).intersection(eligible))
        for group, variables in secondary_groups.items()
    }
    primary_groups = _selected_groups(primary_groups, selected_primary_groups)
    secondary_groups = _selected_groups(secondary_groups, selected_secondary_groups)

    tasks: list[tuple[str, str, list[str]]] = []
    if run_primary_level:
        tasks.extend(("primary", group, variables) for group, variables in primary_groups.items())
    if run_secondary_level:
        tasks.extend(("secondary", group, variables) for group, variables in secondary_groups.items())

    hierarchy_df.to_csv(output_dir / "dictionary_hierarchy_structure.csv", index=False)
    hierarchy_summary.to_csv(output_dir / "dictionary_hierarchy_summary.csv", index=False)
    (output_dir / "hierarchy_tree.txt").write_text(
        hierarchy_tree_text(hierarchy_summary), encoding="utf-8"
    )

    analyses: dict[str, dict[str, Any]] = {"primary": {}, "secondary": {}}
    summary_frames: list[pd.DataFrame] = []
    positive_frames: list[pd.DataFrame] = []
    negative_frames: list[pd.DataFrame] = []
    all_rule_frames: list[pd.DataFrame] = []
    diagnostics_frames: list[pd.DataFrame] = []
    coverage_rows: list[dict[str, Any]] = []

    iterator = tasks
    if config.show_progress_bar and tqdm is not None:
        iterator = tqdm(
            tasks,
            total=len(tasks),
            desc="Primary/secondary Bayes evidence",
            unit="group",
            leave=config.progress_leave,
            dynamic_ncols=True,
            mininterval=config.progress_mininterval,
        )

    for task_index, (level, group, variables) in enumerate(iterator, start=1):
        if config.show_progress_bar and tqdm is not None:
            iterator.set_postfix_str(
                f"{level}: {group} · {len(variables)} variables",
                refresh=False,
            )

        usable = [v for v in variables if v in cluster_df.columns and v not in excluded]
        if len(usable) < config.minimum_variables_per_group:
            coverage_rows.append({
                "level": level,
                "group": group,
                "n_requested_variables": len(variables),
                "n_usable_variables": len(usable),
                "status": "skipped",
                "reason": "insufficient usable variables",
            })
            continue

        group_dir = output_dir / f"{level}_level" / _safe_name(group)
        try:
            analysis = analyze_signed_evidence_group(
                cluster_df=cluster_df,
                dictionary_df=dictionary_df,
                variables=usable,
                cluster_column=cluster_column,
                level=level,
                group=group,
                output_dir=group_dir,
                config=config,
            )
        except ValueError as exc:
            coverage_rows.append({
                "level": level,
                "group": group,
                "n_requested_variables": len(variables),
                "n_usable_variables": len(usable),
                "status": "skipped",
                "reason": str(exc),
            })
            continue

        analyses[level][group] = analysis
        summary_frames.append(analysis["summary"])
        positive_frames.append(analysis["top_positive_rules"])
        negative_frames.append(analysis["top_negative_rules"])
        all_rule_frames.append(analysis["all_rules"])
        diagnostics = analysis["encoding_diagnostics"].copy()
        diagnostics.insert(0, "group", group)
        diagnostics.insert(0, "level", level)
        diagnostics_frames.append(diagnostics)
        coverage_rows.append({
            "level": level,
            "group": group,
            "n_requested_variables": len(variables),
            "n_usable_variables": len(usable),
            "n_encoded_variables": analysis["metadata"]["n_encoded_variables"],
            "n_binary_rules": analysis["metadata"]["n_binary_rules"],
            "n_clusters": analysis["metadata"]["n_clusters"],
            "status": "completed",
            "reason": "",
        })

    summary = pd.concat(summary_frames, ignore_index=True) if summary_frames else pd.DataFrame()
    top_positive = pd.concat(positive_frames, ignore_index=True) if positive_frames else pd.DataFrame()
    top_negative = pd.concat(negative_frames, ignore_index=True) if negative_frames else pd.DataFrame()
    all_rules = pd.concat(all_rule_frames, ignore_index=True) if all_rule_frames else pd.DataFrame()
    diagnostics = pd.concat(diagnostics_frames, ignore_index=True) if diagnostics_frames else pd.DataFrame()
    coverage = pd.DataFrame(coverage_rows)

    summary.to_csv(output_dir / "signed_evidence_all_groups.csv", index=False)
    coverage.to_csv(output_dir / "group_coverage.csv", index=False)
    top_positive.to_csv(output_dir / "top_positive_rules_all_groups.csv", index=False)
    top_negative.to_csv(output_dir / "top_negative_rules_all_groups.csv", index=False)
    if not all_rules.empty:
        cluster_order = sorted(
            all_rules["cluster"].dropna().unique().tolist(),
            key=_cluster_sort_key,
        )
        sorted_rules = all_rules.copy()
        sorted_rules["_cluster_order"] = pd.Categorical(
            sorted_rules["cluster"],
            categories=cluster_order,
            ordered=True,
        )
        sorted_rules = (
            sorted_rules
            .sort_values(
                [
                    "primary_group",
                    "secondary_group",
                    "level",
                    "_cluster_order",
                    "var",
                    "valor_crudo",
                ],
                kind="stable",
            )
            .drop(columns="_cluster_order")
            .reset_index(drop=True)
        )

        master_rules = sorted_rules.loc[:, MASTER_RULE_COLUMNS].copy()
        internal_operations = sorted_rules.loc[:, INTERNAL_OPERATION_COLUMNS].copy()
        count_operations = internal_operations.loc[:, COUNT_OPERATION_COLUMNS].copy()
        score_operations = internal_operations.loc[:, SCORE_OPERATION_COLUMNS].copy()
        epsilon_operations = internal_operations.loc[:, EPSILON_OPERATION_COLUMNS].copy()

        filtered_rules = master_rules.loc[
            master_rules["Nx"].gt(config.export_min_nx)
            & master_rules["n_cluster_x"].gt(config.export_min_n_cluster_x)
            & master_rules["score"].ge(config.export_min_score)
        ].copy()

        master_rules.to_csv(
            output_dir / "bayes_variables_bins_epsilon_score_all_groups.csv",
            index=False,
        )
        filtered_rules.to_csv(
            output_dir / "bayes_variables_bins_epsilon_score_filtered.csv",
            index=False,
        )
    else:
        sorted_rules = pd.DataFrame()
        master_rules = pd.DataFrame(columns=MASTER_RULE_COLUMNS)
        filtered_rules = pd.DataFrame(columns=MASTER_RULE_COLUMNS)
        internal_operations = pd.DataFrame(columns=INTERNAL_OPERATION_COLUMNS)
        count_operations = pd.DataFrame(columns=COUNT_OPERATION_COLUMNS)
        score_operations = pd.DataFrame(columns=SCORE_OPERATION_COLUMNS)
        epsilon_operations = pd.DataFrame(columns=EPSILON_OPERATION_COLUMNS)
    diagnostics.to_csv(output_dir / "encoding_diagnostics_all_groups.csv", index=False)

    heatmap_paths: list[Path] = []
    complex_heatmap_results: dict[str, Any] = {}
    if not summary.empty and config.save_complex_heatmaps:
        from .heatmaps import (
            ComplexHeatmapConfig,
            generate_weighted_complex_heatmaps,
        )

        complex_heatmap_results = generate_weighted_complex_heatmaps(
            summary=summary,
            hierarchy_structure=hierarchy_df,
            output_dir=output_dir / "complex_heatmaps",
            config=ComplexHeatmapConfig(
                publication_dpi=config.publication_dpi,
                robust_percentile=config.complex_heatmap_robust_percentile,
            ),
        )
        heatmap_paths = [
            Path(path)
            for paths in complex_heatmap_results.get("figure_paths", {}).values()
            for path in paths
        ]

    def _write_rules_workbook(
        table: pd.DataFrame,
        path: Path,
        *,
        sheet_name: str,
    ) -> None:
        """Write one audit-friendly evidence table with stable formatting."""
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            table.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            widths = {
                "A": 12,  # level
                "B": 42,  # analysis group
                "C": 36,  # primary questionnaire group
                "D": 46,  # secondary questionnaire group
                "E": 10,  # cluster
                "F": 16,  # var
                "G": 22,  # alias
                "H": 48,  # description
                "I": 58,  # pregunta
                "J": 20,  # variable type
                "K": 34,  # category/bin/range
                "L": 16,  # raw value/bin
                "M": 12,  # Nx
                "N": 13,  # n_cluster
                "O": 14,  # n_cluster_x
                "P": 20,  # coverage_in_cluster
                "Q": 15,  # p_c_given_x
                "R": 14,  # epsilon
                "S": 14,  # score
            }
            for column_letter, width in widths.items():
                worksheet.column_dimensions[column_letter].width = width

            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="1F4E78")
                cell.alignment = cell.alignment.copy(
                    horizontal="center", vertical="center", wrap_text=True
                )

            for cell in worksheet["P"][1:]:
                cell.number_format = "0.0%"
            for cell in worksheet["Q"][1:]:
                cell.number_format = "0.0%"
            for column_letter in ("R", "S"):
                for cell in worksheet[column_letter][1:]:
                    cell.number_format = "0.000"

    def _write_operation_workbook(
        table: pd.DataFrame,
        path: Path,
        *,
        sheet_name: str,
        formula_rows: list[dict[str, str]],
    ) -> None:
        """Write all row-wise count, coverage, score and epsilon operations."""
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            table.to_excel(writer, sheet_name=sheet_name, index=False)
            pd.DataFrame(formula_rows).to_excel(
                writer,
                sheet_name="Formulas",
                index=False,
            )

            data_sheet = writer.book[sheet_name]
            data_sheet.freeze_panes = "A2"
            data_sheet.auto_filter.ref = data_sheet.dimensions
            for cell in data_sheet[1]:
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="44546A")
                cell.alignment = cell.alignment.copy(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Keep identifier columns readable without expensive full-sheet autofit.
            for column_letter, width in {
                "A": 12, "B": 40, "C": 34, "D": 44, "E": 10, "F": 16,
                "G": 20, "H": 42, "I": 52, "J": 18, "K": 30, "L": 16,
            }.items():
                data_sheet.column_dimensions[column_letter].width = width
            from openpyxl.utils import get_column_letter
            for column_index in range(13, data_sheet.max_column + 1):
                data_sheet.column_dimensions[get_column_letter(column_index)].width = 19

            header_to_column = {
                cell.value: get_column_letter(cell.column) for cell in data_sheet[1]
            }
            for field in (
                "coverage_in_cluster", "coverage_in_rest", "prior_cluster",
                "p_x_given_cluster", "p_x_given_rest", "p_c_given_x",
            ):
                column_letter = header_to_column.get(field)
                if column_letter:
                    for cell in data_sheet[column_letter][1:]:
                        cell.number_format = "0.0%"
            for field in (
                "likelihood_ratio", "raw_score", "score",
                "epsilon_difference", "epsilon_numerator",
                "epsilon_denominator", "epsilon",
            ):
                column_letter = header_to_column.get(field)
                if column_letter:
                    for cell in data_sheet[column_letter][1:]:
                        cell.number_format = "0.000000"

            formula_sheet = writer.book["Formulas"]
            formula_sheet.freeze_panes = "A2"
            formula_sheet.auto_filter.ref = formula_sheet.dimensions
            formula_sheet.column_dimensions["A"].width = 34
            formula_sheet.column_dimensions["B"].width = 74
            formula_sheet.column_dimensions["C"].width = 76
            for cell in formula_sheet[1]:
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="44546A")
                cell.alignment = cell.alignment.copy(
                    horizontal="center", vertical="center", wrap_text=True
                )
            for row in formula_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = cell.alignment.copy(vertical="top", wrap_text=True)

    workbook_path = output_dir / "hierarchical_signed_bayes_evidence.xlsx"
    all_variables_workbook_path = (
        output_dir / "hierarchical_signed_bayes_evidence_all_variables.xlsx"
    )
    internal_operations_workbook_path = (
        output_dir / "bayes_internal_operations_all_variables.xlsx"
    )

    _write_rules_workbook(
        filtered_rules,
        workbook_path,
        sheet_name="NB_principal",
    )
    _write_rules_workbook(
        master_rules,
        all_variables_workbook_path,
        sheet_name="NB_all_variables",
    )
    _write_operation_workbook(
        internal_operations,
        internal_operations_workbook_path,
        sheet_name="NB_internal_operations",
        formula_rows=[
            {
                "field": "N",
                "operation": "Total number of analyzed participants.",
                "interpretation": "Global denominator of the analytical sample.",
            },
            {
                "field": "n_cluster / n_rest",
                "operation": "n_cluster = Nc; n_rest = N - n_cluster.",
                "interpretation": "Sizes of the cluster and the one-vs-rest set.",
            },
            {
                "field": "Nx",
                "operation": "Global count of rule X across all participants.",
                "interpretation": "Total support of the category or bin.",
            },
            {
                "field": "n_cluster_x / n_rest_x",
                "operation": "Counts of X inside and outside the cluster.",
                "interpretation": "Contingency table used for coverage, score, and epsilon.",
            },
            {
                "field": "coverage_in_cluster",
                "operation": "n_cluster_x / n_cluster",
                "interpretation": "Proportion of cluster members presenting X.",
            },
            {
                "field": "coverage_in_rest",
                "operation": "n_rest_x / n_rest",
                "interpretation": "Proportion of the remaining participants presenting X.",
            },
            {
                "field": "prior_cluster",
                "operation": "n_cluster / N",
                "interpretation": "Baseline cluster prevalence in the analytical sample.",
            },
            {
                "field": "p_x_given_cluster",
                "operation": "(n_cluster_x + alpha) / (n_cluster + 2*alpha)",
                "interpretation": "Smoothed probability of observing X inside the cluster.",
            },
            {
                "field": "p_x_given_rest",
                "operation": "(n_rest_x + alpha) / (n_rest + 2*alpha)",
                "interpretation": "Smoothed probability of observing X outside the cluster.",
            },
            {
                "field": "likelihood_ratio",
                "operation": "p_x_given_cluster / p_x_given_rest",
                "interpretation": "Cluster likelihood ratio relative to the rest.",
            },
            {
                "field": "raw_score",
                "operation": "ln(likelihood_ratio)",
                "interpretation": "Bayes score before applying the minimum-support rule.",
            },
            {
                "field": "score_aplicado",
                "operation": "1 when Nx >= min_cases; otherwise 0.",
                "interpretation": "Indicates whether the rule has sufficient support.",
            },
            {
                "field": "score",
                "operation": "raw_score when score_aplicado = 1; otherwise 0.",
                "interpretation": "Final evidence: positive favors the cluster and negative disfavors it.",
            },
            {
                "field": "p_c_given_x",
                "operation": "(n_cluster_x + alpha) / (Nx + 2*alpha)",
                "interpretation": "Smoothed probability of cluster membership among participants presenting X.",
            },
            {
                "field": "epsilon_difference",
                "operation": "p_c_given_x - prior_cluster",
                "interpretation": "Change relative to the baseline cluster prevalence.",
            },
            {
                "field": "epsilon_numerator",
                "operation": "Nx * epsilon_difference",
                "interpretation": "Deviation weighted by the global support of X.",
            },
            {
                "field": "epsilon_denominator",
                "operation": "sqrt(Nx * prior_cluster * (1 - prior_cluster))",
                "interpretation": "Binomial scale used to standardize the deviation.",
            },
            {
                "field": "epsilon",
                "operation": "epsilon_numerator / epsilon_denominator",
                "interpretation": "Standardized deviation of P(C|X) relative to P(C).",
            },
        ],
    )

    manifest = {
        "analysis_scope": ANALYSIS_SCOPE,
        "predictive_metrics_computed": False,
        "dataset_split": False,
        "run_primary_level": run_primary_level,
        "run_secondary_level": run_secondary_level,
        "n_groups_requested": len(tasks),
        "n_groups_completed": int(coverage["status"].eq("completed").sum()) if not coverage.empty else 0,
        "n_groups_skipped": int(coverage["status"].eq("skipped").sum()) if not coverage.empty else 0,
        "visualization_measure": "weighted evidence per effective variable",
        "heatmap_engine": "PyComplexHeatmap",
        "filtered_export_rule": {
            "Nx": f"> {config.export_min_nx}",
            "n_cluster_x": f"> {config.export_min_n_cluster_x}",
            "score": f">= {config.export_min_score}",
        },
        "n_rules_all_variables": int(len(master_rules)),
        "n_rules_filtered": int(len(filtered_rules)),
        "internal_operations_workbook": str(internal_operations_workbook_path),
        "config": asdict(config),
    }
    (output_dir / "run_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    return {
        "hierarchy_structure": hierarchy_df,
        "hierarchy_summary": hierarchy_summary,
        "hierarchy_tree": hierarchy_tree_text(hierarchy_summary),
        "coverage": coverage,
        "summary": summary,
        "top_positive_rules": top_positive,
        "top_negative_rules": top_negative,
        "all_rules": all_rules,
        "master_rules_table": master_rules,
        "filtered_rules_table": filtered_rules,
        "internal_operations_table": internal_operations,
        "count_operations_table": count_operations,
        "score_operations_table": score_operations,
        "epsilon_operations_table": epsilon_operations,
        "encoding_diagnostics": diagnostics,
        "analyses": analyses,
        "workbook_path": workbook_path,
        "all_variables_workbook_path": all_variables_workbook_path,
        "internal_operations_workbook_path": internal_operations_workbook_path,
        "heatmap_paths": heatmap_paths,
        "complex_heatmap_results": complex_heatmap_results,
        "manifest": manifest,
    }
